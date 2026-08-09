"""Импорт каталога из выгрузок площадок (CSV/XLSX).

Логика в два шага:
1. Загружаем файл → показываем его колонки и просим сопоставить с полями книги.
2. По сопоставлению создаём/обновляем книги. Сопоставление одинаковых книг между
   площадками идёт по SKU или ISBN — если совпало, дополняем существующую книгу
   лотом нужной площадки, а не плодим дубли.
"""
import csv
import io
import secrets
import threading
import time
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.catalog_sync import TARGET_FIELDS, sync_all, sync_marketplace, upsert_catalog_rows
from app.db import get_db
from app.marketplaces import MarketplaceError, is_supported
from app.models import Marketplace, MarketplaceAccount, SyncLog
from app.templating import templates

router = APIRouter(prefix="/import")

# Словарь автосопоставления: какие заголовки колонок в выгрузках Ozon/WB
# соответствуют полям книги. Сравниваем по подстроке в нижнем регистре, поэтому
# хватает характерных кусков названий ("артикул", "цена", "штрихкод" и т.д.).
# Порядок важен: первое совпадение выигрывает.
COLUMN_ALIASES = {
    "sku": ["артикул продавца", "артикул", "offer_id", "sku", "ваш sku", "код товара"],
    "title": ["название товара", "наименование", "название", "заголовок", "title", "name"],
    "author": [
        "автор", "авторы", "author", "автор книги", "автор(ы)",
        "составитель", "писатель", "author name", "авт.",
    ],
    "isbn": ["isbn", "штрихкод", "штрих-код", "barcode", "ean"],
    "publisher": ["издательство", "бренд", "publisher", "brand"],
    "year": ["год выпуска", "год издания", "год", "year"],
    "condition": ["состояние", "качество", "condition"],
    "price": ["цена продажи", "текущая цена", "цена", "price"],
    "description": ["описание", "аннотация", "description"],
    "external_id": [
        "ozon product id", "product id", "product_id", "id товара",
        "ozon id", "sku ozon", "артикул ozon",
    ],
}

# Автоопределение площадки по характерным колонкам выгрузки.
MARKETPLACE_HINTS = {
    "ozon": ["ozon", "offer_id", "артикул ozon", "fbo", "fbs"],
    "wildberries": ["wildberries", "wb", "номенклатура", "предмет", "баркод"],
}


def _auto_map(columns: list[str]) -> dict[str, str]:
    """Подобрать колонку под каждое поле книги по словарю синонимов.

    Возвращает {поле: имя_колонки}. Одну колонку не назначаем двум полям.
    """
    lowered = {c: (c or "").strip().lower() for c in columns}
    mapping: dict[str, str] = {}
    used: set[str] = set()
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            match = next(
                (c for c in columns if c not in used and alias in lowered[c]), None
            )
            if match:
                mapping[field] = match
                used.add(match)
                break
    return mapping


def _guess_marketplace(columns: list[str], filename: str) -> str | None:
    """Угадать площадку по колонкам и имени файла. None — если непонятно."""
    haystack = " ".join(columns).lower() + " " + (filename or "").lower()
    for mp, hints in MARKETPLACE_HINTS.items():
        if any(h in haystack for h in hints):
            return mp
    return None


# Простое хранилище загруженного файла между шагом 1 и шагом 2 (по сессии).
# Ключ — случайный токен (secrets.token_urlsafe), чтобы избежать коллизий между
# пользователями и предотвратить утечку данных. Записи старше 10 минут автоудаляются.
# Блокировка защищает от race condition при параллельных загрузках.
_uploads: dict[str, tuple[list[dict], float]] = {}  # {token: (rows, timestamp)}
_uploads_lock = threading.Lock()


def _cleanup_old_uploads():
    """Удалить загрузки старше 10 минут, чтобы избежать утечки памяти."""
    now = time.time()
    with _uploads_lock:
        expired = [token for token, (_, ts) in _uploads.items() if now - ts > 600]
        for token in expired:
            _uploads.pop(token, None)


def _parse_file(filename: str, raw: bytes) -> list[dict]:
    """Читаем CSV или XLSX в список словарей {колонка: значение}."""
    name = filename.lower()
    if name.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="replace")
        # Пытаемся угадать разделитель (Ozon/WB часто отдают ; )
        sample = text[:5000]
        delimiter = ";" if sample.count(";") > sample.count(",") else ","
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        return [dict(row) for row in reader]
    if name.endswith(".xls") and not name.endswith(".xlsx"):
        raise ValueError(
            "Старый формат .xls не поддерживается. Откройте файл в Excel и "
            "сохраните как .xlsx (или экспортируйте в CSV)."
        )
    if name.endswith(".xlsx"):
        from openpyxl import load_workbook

        # Намеренно НЕ используем read_only: у выгрузок Ozon/1С часто указан
        # неверный размер листа, из-за чего быстрый парсер читает лишь первую
        # строку. Обычный режим читает лист целиком, пусть и медленнее.
        try:
            wb = load_workbook(io.BytesIO(raw), data_only=True)
        except Exception as exc:
            raise ValueError(f"Не удалось открыть файл Excel: {exc}") from exc
        ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)

        def nonempty(row):
            return [c for c in row if c is not None and str(c).strip() != ""]

        # Ищем строку-шапку: первую, где заполнено хотя бы две ячейки. Так
        # пропускаем титульные строки отчёта ("Отчёт по товарам" в одной ячейке).
        headers = None
        for row in rows_iter:
            if len(nonempty(row)) >= 2:
                headers = [str(h).strip() if h is not None else "" for h in row]
                break
        if not headers:
            return []

        result = []
        for row in rows_iter:
            if not nonempty(row):  # пропускаем полностью пустые строки
                continue
            result.append(
                {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
            )
        return result
    raise ValueError("Поддерживаются только CSV и XLSX")


@router.get("", response_class=HTMLResponse)
def import_start(request: Request, db: Session = Depends(get_db)):
    # Кнопки «Загрузить из ...» активны только для площадок с включёнными ключами.
    return templates.TemplateResponse(
        request,
        "import_start.html",
        {"marketplaces": list(Marketplace), "sources": _sources(db)},
    )


@router.post("/pull/{marketplace}", response_class=HTMLResponse)
def import_pull(marketplace: str, request: Request, db: Session = Depends(get_db)):
    """Полная сверка каталога одной площадки по сохранённым ключам.

    Тянет каталог, апсертит книги и снимает пропавшее (кросс-снятие). Это та же
    операция, что и фоновая сверка, но запущенная вручную по конкретной площадке.
    """
    def fail(msg: str):
        db.add(SyncLog(marketplace=marketplace, action="import_pull", ok=False, message=msg))
        db.commit()
        return templates.TemplateResponse(
            request,
            "import_start.html",
            {"marketplaces": list(Marketplace), "sources": _sources(db), "error": msg},
            status_code=400,
        )

    if not is_supported(marketplace):
        return fail(f"Площадка «{marketplace}» не поддерживает загрузку по API")

    try:
        result = sync_marketplace(db, marketplace)
        db.commit()
    except MarketplaceError as exc:
        return fail(f"Не удалось сверить каталог: {exc}")
    except Exception as exc:  # noqa: BLE001 — любой сбой показываем как есть
        db.rollback()
        return fail(f"Ошибка сверки: {exc}")

    return templates.TemplateResponse(
        request,
        "import_done.html",
        {**result, "marketplace": marketplace, "auto": True, "via_api": True},
    )


@router.post("/sync")
def import_sync(request: Request, db: Session = Depends(get_db)):
    """Кнопка «Обновить каталог»: полная сверка всех включённых площадок сразу."""
    results = sync_all(db)
    if not results:
        return RedirectResponse("/?synced=" + quote("нет включённых площадок"), status_code=303)
    if "__busy__" in results:
        return RedirectResponse("/?synced=" + quote("Сверка уже выполняется"), status_code=303)

    parts = []
    for mp, res in results.items():
        if "error" in res:
            parts.append(f"{mp}: ошибка")
        else:
            parts.append(
                f"{mp}: +{res['created']} новых, {res['updated']} обновлено, "
                f"{res['removed']} снято"
            )
    return RedirectResponse("/?synced=" + quote("; ".join(parts)), status_code=303)


@router.post("/fix-wb-ids")
def fix_wb_external_ids(request: Request, db: Session = Depends(get_db)):
    """Служебная ручка: обновить external_id у всех WB-лотов на актуальный nmID.

    Запускается вручную один раз для миграции старых данных (где external_id был
    vendorCode). После этого catalog_sync.py автоматически обновляет external_id
    при каждой сверке, поэтому повторный запуск не нужен.
    """
    from app.access import require_section
    from app.marketplaces import get_client
    from app.models import Listing, MarketplaceAccount
    from app.security import decrypt_credentials

    # Только для владельца или руководителя (доступ к настройкам)
    require_section(request, "settings")

    # Проверяем настройки WB
    account = db.scalar(
        select(MarketplaceAccount).where(MarketplaceAccount.marketplace == "wildberries")
    )
    if not account or not account.enabled or not account.credentials_encrypted:
        return RedirectResponse("/settings?error=" + quote("WB площадка выключена или нет ключей"), status_code=303)

    try:
        creds = decrypt_credentials(account.credentials_encrypted)
        client = get_client("wildberries", creds)
    except Exception as exc:
        return RedirectResponse("/settings?error=" + quote(f"Не удалось подключиться к WB: {exc}"), status_code=303)

    # Находим все WB-лоты
    listings = db.scalars(
        select(Listing)
        .options(selectinload(Listing.book))
        .where(Listing.marketplace == "wildberries")
    ).all()

    if not listings:
        return RedirectResponse("/settings?error=" + quote("WB-лотов в базе нет"), status_code=303)

    # Запрашиваем все карточки с WB
    try:
        cards_data = client.list_catalog()
    except MarketplaceError as exc:
        return RedirectResponse("/settings?error=" + quote(f"Не удалось получить каталог WB: {exc}"), status_code=303)

    # Индекс по vendorCode для быстрого поиска
    cards_by_vendor = {}
    for card in cards_data:
        vendor_code = card.get("sku")  # list_catalog возвращает sku как vendorCode
        nm_id = card.get("external_id")  # и external_id как nmID (если есть)
        if vendor_code and nm_id:
            try:
                # Проверяем, что nmID — это число
                int(nm_id)
                cards_by_vendor[vendor_code] = nm_id
            except (ValueError, TypeError):
                pass

    updated = 0
    already_ok = 0

    for listing in listings:
        current_ext_id = listing.external_id

        # Проверяем, нужно ли обновлять
        if current_ext_id:
            try:
                int(current_ext_id)
                # Уже число — всё ок
                already_ok += 1
                continue
            except (ValueError, TypeError):
                pass  # Не число — нужно обновить

        # Ищем nmID по SKU
        sku = listing.sku
        if not sku:
            continue

        nm_id = cards_by_vendor.get(sku)
        if nm_id and nm_id != current_ext_id:
            listing.external_id = nm_id
            updated += 1

    db.commit()

    msg = f"WB external_id обновлён: {updated} лотов, {already_ok} уже корректных"
    return RedirectResponse("/settings?wb_ids=" + quote(msg), status_code=303)


def _sources(db: Session) -> list[dict]:
    """Список площадок с признаком готовности (ключи включены) для шаблона."""
    accounts = {a.marketplace: a for a in db.scalars(select(MarketplaceAccount)).all()}
    out = []
    for mp in Marketplace:
        acc = accounts.get(mp.value)
        out.append(
            {
                "marketplace": mp.value,
                "ready": bool(
                    is_supported(mp.value) and acc and acc.enabled and acc.credentials_encrypted
                ),
            }
        )
    return out


MAX_UPLOAD_BYTES = 50 * 1024 * 1024  # 50 МБ — достаточно для любой выгрузки Ozon/WB


@router.post("/upload", response_class=HTMLResponse)
async def import_upload(
    request: Request,
    db: Session = Depends(get_db),
    file: UploadFile = File(...),
    marketplace: str = Form(""),
):
    """Загрузка выгрузки. Площадку и сопоставление колонок определяем сами.

    Если удалось распознать SKU/название — импортируем сразу, без лишних шагов.
    Показываем экран сопоставления только когда автоопределение не справилось.
    """
    # Проверяем размер ДО чтения в память: защита от DoS через огромные файлы
    content_length = request.headers.get("content-length")
    if content_length and int(content_length) > MAX_UPLOAD_BYTES:
        return templates.TemplateResponse(
            request,
            "import_start.html",
            {"marketplaces": list(Marketplace), "sources": _sources(db),
             "error": f"Файл слишком большой (максимум {MAX_UPLOAD_BYTES // 1024 // 1024} МБ)"},
            status_code=413,
        )

    raw = await file.read(MAX_UPLOAD_BYTES + 1)
    if len(raw) > MAX_UPLOAD_BYTES:
        return templates.TemplateResponse(
            request,
            "import_start.html",
            {"marketplaces": list(Marketplace), "sources": _sources(db),
             "error": f"Файл слишком большой (максимум {MAX_UPLOAD_BYTES // 1024 // 1024} МБ)"},
            status_code=413,
        )
    try:
        rows = _parse_file(file.filename or "", raw)
    except Exception as exc:
        return templates.TemplateResponse(
            request,
            "import_start.html",
            {"marketplaces": list(Marketplace), "sources": _sources(db), "error": str(exc)},
            status_code=400,
        )

    if not rows:
        return templates.TemplateResponse(
            request,
            "import_start.html",
            {"marketplaces": list(Marketplace), "sources": _sources(db),
             "error": "Файл пустой"},
            status_code=400,
        )

    columns = list(rows[0].keys())

    # Площадку берём из формы, если выбрали вручную, иначе угадываем по файлу.
    if not marketplace:
        marketplace = _guess_marketplace(columns, file.filename or "") or "ozon"

    # Токен случайный, а не «площадка:имя файла»: одинаковое имя файла у двух
    # пользователей раньше перетирало чужую загрузку. Заодно чистим старые записи,
    # чтобы брошенные (не доведённые до шага 2) выгрузки не висели в памяти.
    _cleanup_old_uploads()
    token = secrets.token_urlsafe(16)
    with _uploads_lock:
        _uploads[token] = (rows, time.time())
    request.session["import_token"] = token
    request.session["import_marketplace"] = marketplace

    # Пробуем полностью автоматический импорт.
    auto = _auto_map(columns)
    if auto.get("sku") or auto.get("title"):
        result = _do_import(db, marketplace, rows, auto)
        with _uploads_lock:
            _uploads.pop(token, None)
        return templates.TemplateResponse(
            request,
            "import_done.html",
            {**result, "marketplace": marketplace, "auto": True},
        )

    # Автоопределение не нашло даже название/артикул — просим сопоставить вручную.
    return templates.TemplateResponse(
        request,
        "import_map.html",
        {
            "columns": columns,
            "target_fields": TARGET_FIELDS,
            "sample": rows[:5],
            "marketplace": marketplace,
            "total": len(rows),
            "auto_map": auto,
        },
    )


@router.post("/run")
async def import_run(request: Request, db: Session = Depends(get_db)):
    form = await request.form()
    token = request.session.get("import_token")
    marketplace = request.session.get("import_marketplace")
    # В _uploads лежит кортеж (строки, время загрузки) — распаковываем, иначе в
    # импорт ушёл бы сам кортеж. Пустой токен или истёкшая запись → на шаг 1.
    with _uploads_lock:
        entry = _uploads.get(token) if token else None
        rows = entry[0] if entry else None
    if not rows or not marketplace:
        return RedirectResponse("/import", status_code=303)

    # mapping: поле_книги -> имя_колонки_в_файле
    mapping = {
        field: form.get(f"map_{field}", "")
        for field in TARGET_FIELDS
        if form.get(f"map_{field}")
    }

    result = _do_import(db, marketplace, rows, mapping)
    with _uploads_lock:
        _uploads.pop(token, None)
    return templates.TemplateResponse(
        request,
        "import_done.html",
        {**result, "marketplace": marketplace},
    )


def _do_import(db: Session, marketplace: str, rows: list[dict], mapping: dict) -> dict:
    """Импорт из файла: апсерт книг без сверки пропавших.

    В отличие от полной сверки по API, файл может быть частичной выгрузкой,
    поэтому НЕ снимаем книги, которых в файле нет (reconcile не запускаем).
    Общий апсерт живёт в catalog_sync.upsert_catalog_rows.
    """
    result = upsert_catalog_rows(db, marketplace, rows, mapping)
    db.add(
        SyncLog(
            marketplace=marketplace,
            action="import",
            ok=True,
            message=(f"Импорт файлом: создано {result['created']}, "
                     f"обновлено {result['updated']}, пропущено {result['skipped']}"),
        )
    )
    db.commit()
    return {"created": result["created"], "updated": result["updated"], "skipped": result["skipped"]}
